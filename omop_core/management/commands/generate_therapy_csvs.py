"""Parse therapy spreadsheet and generate seed CSVs with Athena concept resolution.

Usage:
    DATABASE_URL="${STAGING_DATABASE_URL:-$DATABASE_URL}" \\
      .venv/bin/python manage.py generate_therapy_csvs \\
        --input ~/Downloads/Therapies_Therapy\\ Comp_Therapy\\ Types.xlsx
"""
import csv
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

import logging

logger = logging.getLogger(__name__)


def _slugify(text):
    """Convert display title to a code slug."""
    s = text.strip().lower()
    s = re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_')


_ROUND_MAP = {
    'first line therapy': 'first_line_therapy',
    'second line therapy': 'second_line_therapy',
    'later therapy': 'later_line_therapy',
    'supportive therapy': 'supportive_therapy',
    # Alternate spellings
    'first-line therapy': 'first_line_therapy',
    'second-line therapy': 'second_line_therapy',
    'later-line therapy': 'later_line_therapy',
    '1st line therapy': 'first_line_therapy',
    '2nd line therapy': 'second_line_therapy',
    'secondline therapy': 'second_line_therapy',
    'second round therapy': 'second_line_therapy',
}

# Disease names → Disease.code values.  Must match the codes in the Disease
# vocabulary table (seeded by migrations 0056/0057 and 0178).
_DISEASE_MAP = {
    'mantle cell lymphoma': 'MCL',
    'chronic lymphocytic leukemia': 'C2987',
    'multiple myeloma': 'C3242',
    'follicular lymphoma': 'C3209',
    'breast cancer': 'C9335',
    'diffuse large b-cell lymphoma': 'DLBCL',
}


class Command(BaseCommand):
    help = 'Parse therapy spreadsheet and generate seed CSVs with Athena concept_ids.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--input', required=True,
            help='Path to the therapy Excel file (.xlsx)',
        )
        parser.add_argument(
            '--output-dir', default='data',
            help='Directory to write CSVs into (default: data/)',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                'openpyxl is required. Install it: pip install openpyxl'
            )

        xlsx_path = Path(options['input']).expanduser()
        if not xlsx_path.exists():
            raise CommandError(f'File not found: {xlsx_path}')

        out_dir = Path(options['output_dir'])
        out_dir.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb[wb.sheetnames[0]]

        # ── Parse the spreadsheet ──────────────────────────────────────────
        therapies = {}          # code → title
        components = {}         # code → title
        classes = {}            # code → title
        regimen_components = set()   # (therapy_code, component_code)
        component_classes = set()    # (component_code, class_code)
        disease_rounds = []          # (therapy_code, disease, round)

        cur_therapy_code = None
        cur_therapy_title = None
        cur_component_code = None
        cur_component_title = None
        cur_disease = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= 1:  # skip headers
                continue

            t_code, t_title, disease, rnd, c_code, c_title, cl_code, cl_title = (
                (str(v).strip() if v else None) for v in (list(row) + [None] * 8)[:8]
            )

            # Strip leading/trailing whitespace and tabs from titles
            if t_title:
                t_title = t_title.strip().strip('\t')

            # Carry-forward: therapy
            if t_code:
                cur_therapy_code = t_code.strip()
                cur_therapy_title = t_title or cur_therapy_code
                therapies[cur_therapy_code] = cur_therapy_title
                cur_component_code = None
                cur_component_title = None
                cur_disease = None

            # Carry-forward: disease
            if disease:
                cur_disease = disease.strip()

            # Disease-round linkage
            if rnd and cur_therapy_code:
                rnd_clean = rnd.strip().lower()
                round_code = _ROUND_MAP.get(rnd_clean)
                if round_code is None:
                    logger.warning(f'Row {i+1}: unrecognized round "{rnd}"')
                    round_code = _slugify(rnd)

                disease_code = None
                if cur_disease:
                    disease_code = _DISEASE_MAP.get(cur_disease.lower())
                    if disease_code is None:
                        logger.warning(f'Row {i+1}: unrecognized disease "{cur_disease}"')
                        disease_code = _slugify(cur_disease)

                if disease_code:
                    disease_rounds.append((cur_therapy_code, disease_code, round_code))

            # Carry-forward: component
            if c_code:
                cur_component_code = c_code.strip()
                cur_component_title = (c_title or cur_component_code).strip()
                components[cur_component_code] = cur_component_title
                if cur_therapy_code:
                    regimen_components.add((cur_therapy_code, cur_component_code))

            # Classes (always present per row when there's data)
            if cl_code and cur_component_code:
                cl_code_clean = cl_code.strip()
                cl_title_clean = (cl_title or cl_code_clean).strip()
                classes[cl_code_clean] = cl_title_clean
                component_classes.add((cur_component_code, cl_code_clean))

        wb.close()

        self.stdout.write(
            f'Parsed: {len(therapies)} therapies, {len(components)} components, '
            f'{len(classes)} classes, {len(regimen_components)} regimen-component links, '
            f'{len(component_classes)} component-class links, '
            f'{len(disease_rounds)} disease-round links'
        )

        # ── Resolve Athena concept_ids ─────────────────────────────────────
        from omop_core.models import Concept

        def _resolve_hemonc(names, concept_class_id):
            """Resolve a dict of {code: title} to {code: concept_id} using HemOnc."""
            if not names:
                return {}
            result = {}
            # Build a case-insensitive name→concept_id lookup
            title_list = list(set(names.values()))
            qs = Concept.objects.filter(
                vocabulary_id='HemOnc',
                concept_class_id=concept_class_id,
                concept_name__in=title_list,
            ).values_list('concept_name', 'concept_id')
            exact = {n: cid for n, cid in qs}

            # Case-insensitive fallback
            unresolved_titles = [t for t in title_list if t not in exact]
            ilike = {}
            if unresolved_titles:
                for t in unresolved_titles:
                    match = Concept.objects.filter(
                        vocabulary_id='HemOnc',
                        concept_class_id=concept_class_id,
                        concept_name__iexact=t,
                    ).values_list('concept_id', flat=True).first()
                    if match:
                        ilike[t] = match

            for code, title in names.items():
                cid = exact.get(title) or ilike.get(title)
                result[code] = cid
            return result

        def _resolve_rxnorm_ingredient(names):
            """Fallback: resolve component names to RxNorm Ingredient."""
            result = {}
            for code, title in names.items():
                match = Concept.objects.filter(
                    vocabulary_id='RxNorm',
                    concept_class_id='Ingredient',
                    concept_name__iexact=title,
                ).values_list('concept_id', flat=True).first()
                if match:
                    result[code] = match
            return result

        therapy_concepts = _resolve_hemonc(therapies, 'Regimen')
        component_concepts = _resolve_hemonc(components, 'Component')
        class_concepts = _resolve_hemonc(classes, 'Component Class')

        # Fallback: try RxNorm Ingredient for unresolved components
        unresolved_components = {
            code: title for code, title in components.items()
            if component_concepts.get(code) is None
        }
        if unresolved_components:
            rxnorm_hits = _resolve_rxnorm_ingredient(unresolved_components)
            for code, cid in rxnorm_hits.items():
                component_concepts[code] = cid

        # ── Print resolution summary ──────────────────────────────────────
        def _summary(label, items, concepts):
            resolved = sum(1 for c in items if concepts.get(c))
            total = len(items)
            self.stdout.write(f'  {label}: {resolved}/{total} resolved')
            unresolved = [c for c in items if not concepts.get(c)]
            if unresolved:
                for u in unresolved[:20]:
                    self.stdout.write(f'    UNRESOLVED: {u} ({items[u]})')
                if len(unresolved) > 20:
                    self.stdout.write(f'    ... and {len(unresolved) - 20} more')

        self.stdout.write('\nConcept resolution:')
        _summary('Therapies (HemOnc Regimen)', therapies, therapy_concepts)
        _summary('Components (HemOnc Component + RxNorm)', components, component_concepts)
        _summary('Classes (HemOnc Component Class)', classes, class_concepts)

        # ── Write CSVs ────────────────────────────────────────────────────
        tc_path = out_dir / 'therapies_and_components.csv'
        with open(tc_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow([
                'therapy_code', 'therapy_title', 'therapy_concept_id',
                'component_code', 'component_title', 'component_concept_id',
            ])
            for t_code, c_code in sorted(regimen_components):
                w.writerow([
                    t_code, therapies[t_code], therapy_concepts.get(t_code, ''),
                    c_code, components[c_code], component_concepts.get(c_code, ''),
                ])
        self.stdout.write(f'\nWrote {tc_path} ({len(regimen_components)} rows)')

        cc_path = out_dir / 'components_and_classes.csv'
        with open(cc_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow([
                'component_code', 'component_title', 'component_concept_id',
                'class_code', 'class_title', 'class_concept_id',
            ])
            for c_code, cl_code in sorted(component_classes):
                w.writerow([
                    c_code, components[c_code], component_concepts.get(c_code, ''),
                    cl_code, classes[cl_code], class_concepts.get(cl_code, ''),
                ])
        self.stdout.write(f'Wrote {cc_path} ({len(component_classes)} rows)')

        dtr_path = out_dir / 'disease_therapy_rounds.csv'
        with open(dtr_path, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['therapy_code', 'disease', 'round'])
            # Deduplicate
            seen = set()
            for t_code, disease, rnd in disease_rounds:
                key = (t_code, disease, rnd)
                if key not in seen:
                    seen.add(key)
                    w.writerow([t_code, disease, rnd])
        self.stdout.write(f'Wrote {dtr_path} ({len(seen)} rows)')
