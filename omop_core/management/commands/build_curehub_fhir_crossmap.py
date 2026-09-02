"""Build a non-overlapping CureHub FHIR crossmap from the OMOP Mapping Worklist.

Reads the CureHub OMOP Mapping Worklist Excel file and generates
``docs/ht-fhir-code-concept-mapping.md`` — a companion to the existing
``docs/ht-code-concept-mapping.md`` — containing **only** source codes that
do not already appear in the HT-One/Next crossmap.

Sheets processed:
  A — standard vocabulary codes (LOINC, ICD10CM, SNOMED, RxNorm, etc.)
  B — local-namespace codes with a ``shadow_vocab`` column
  C — institutional / PatientRecord-priority codes
  F — top 2,000 other clinical codes

Sheets skipped:
  D — noise / metadata
  E — blocked namespaces (no display text)
"""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError


DEFAULT_HT_CROSSMAP = Path(__file__).resolve().parents[3] / 'docs' / 'ht-code-concept-mapping.md'
DEFAULT_OUTPUT = Path(__file__).resolve().parents[3] / 'docs' / 'ht-fhir-code-concept-mapping.md'

# Map Sheet A vocabulary names to the OMOP vocabulary_id used in the crossmap
# format.  Sheet A already uses OMOP-standard names.
_SHEET_A_VOCAB_MAP = {
    'LOINC': 'LOINC',
    'ICD10CM': 'ICD10CM',
    'SNOMED': 'SNOMED',
    'RxNorm': 'RxNorm',
    'ICD9CM': 'ICD9CM',
    'CVX': 'CVX',
    'CPT4': 'CPT4',
}


def _read_ht_crossmap_keys(path):
    """Return the set of ``(source_vocabulary_id, source_code)`` pairs in the
    existing HT-One/Next crossmap Markdown file."""
    keys = set()
    try:
        text = path.read_text()
    except FileNotFoundError:
        return keys
    for line in text.splitlines():
        if not line.startswith('| ') or line.startswith('| ---') or 'Source system' in line:
            continue
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) < 2:
            continue
        source_vocab = cells[0]
        source_code = cells[1].replace('\\|', '|')
        keys.add((source_vocab, source_code))
    return keys


def _escape_pipe(text):
    """Escape literal pipe characters for Markdown table cells."""
    if text is None:
        return ''
    return str(text).replace('|', '\\|')


class Command(BaseCommand):
    help = 'Build ht-fhir-code-concept-mapping.md from the CureHub OMOP Mapping Worklist.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--input', required=True,
            help='Path to CureHub_OMOP_Mapping_Worklist.xlsx.',
        )
        parser.add_argument(
            '--ht-crossmap', default=str(DEFAULT_HT_CROSSMAP),
            help='Path to existing HT-One/Next crossmap Markdown (for overlap exclusion).',
        )
        parser.add_argument(
            '--output', default=str(DEFAULT_OUTPUT),
            help='Output Markdown path.',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Report counts without writing the output file.',
        )

    def handle(self, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError('openpyxl is required: pip install openpyxl') from exc

        input_path = Path(options['input'])
        if not input_path.exists():
            raise CommandError(f'Input file not found: {input_path}')

        ht_crossmap_path = Path(options['ht_crossmap'])
        self.stdout.write(f'Reading existing crossmap from {ht_crossmap_path} …')
        existing_keys = _read_ht_crossmap_keys(ht_crossmap_path)
        self.stdout.write(f'  {len(existing_keys):,} existing (source_vocab, source_code) pairs loaded.')

        self.stdout.write(f'Reading worklist from {input_path} …')
        wb = openpyxl.load_workbook(str(input_path), read_only=True)

        rows = []
        stats = {
            'A_total': 0, 'A_overlap': 0, 'A_added': 0,
            'B_total': 0, 'B_overlap': 0, 'B_added': 0,
            'C_total': 0, 'C_overlap': 0, 'C_added': 0,
            'F_total': 0, 'F_overlap': 0, 'F_added': 0,
        }
        seen = set()  # track (vocab, code) within this run to avoid intra-sheet dupes

        # --- Sheet A: standard vocabulary codes ---
        if 'A_Standard_verify' in wb.sheetnames:
            ws = wb['A_Standard_verify']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code, vocabulary, text, occurrences, resource_types = (
                    row[0], row[1], row[2], row[3], row[4],
                )
                if code is None or vocabulary is None:
                    continue
                stats['A_total'] += 1
                vocab_id = _SHEET_A_VOCAB_MAP.get(str(vocabulary).strip(), str(vocabulary).strip())
                code_str = str(code).strip()
                key = (vocab_id, code_str)
                if key in existing_keys or key in seen:
                    stats['A_overlap'] += 1
                    continue
                seen.add(key)
                stats['A_added'] += 1
                rows.append({
                    'source_vocab': vocab_id,
                    'source_code': code_str,
                    'target_vocab': vocab_id,   # standard code → same vocab is the target
                    'target_code': code_str,
                    'domain': '',
                    'status': 'proposed',
                    'origins': 'HT-FHIR',
                    'description': str(text or ''),
                    'occurrences': occurrences or 0,
                })

        # --- Sheet B: standard codes in local namespaces ---
        if 'B_Standard_in_local_ns' in wb.sheetnames:
            ws = wb['B_Standard_in_local_ns']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code, vocabulary, text, occurrences, resource_types, shadow_vocab = (
                    row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None,
                )
                if code is None or vocabulary is None:
                    continue
                stats['B_total'] += 1
                vocab_id = str(vocabulary).strip()
                code_str = str(code).strip()
                key = (vocab_id, code_str)
                if key in existing_keys or key in seen:
                    stats['B_overlap'] += 1
                    continue
                # Also check if the shadow vocab + code overlaps
                shadow = str(shadow_vocab).strip() if shadow_vocab else ''
                if shadow and (shadow, code_str) in existing_keys:
                    stats['B_overlap'] += 1
                    continue
                seen.add(key)
                stats['B_added'] += 1
                target_vocab = shadow if shadow and shadow != 'None' else ''
                rows.append({
                    'source_vocab': vocab_id,
                    'source_code': code_str,
                    'target_vocab': target_vocab,
                    'target_code': code_str if target_vocab else '',
                    'domain': '',
                    'status': 'proposed',
                    'origins': 'HT-FHIR',
                    'description': str(text or ''),
                    'occurrences': occurrences or 0,
                })

        # --- Sheet C: PatientRecord priority codes ---
        if 'C_Priority_PatientRecord' in wb.sheetnames:
            ws = wb['C_Priority_PatientRecord']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code, vocabulary, text, occurrences, resource_types, pr_category = (
                    row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None,
                )
                if code is None or vocabulary is None:
                    continue
                stats['C_total'] += 1
                vocab_id = str(vocabulary).strip()
                code_str = str(code).strip()
                key = (vocab_id, code_str)
                if key in existing_keys or key in seen:
                    stats['C_overlap'] += 1
                    continue
                seen.add(key)
                stats['C_added'] += 1
                rows.append({
                    'source_vocab': vocab_id,
                    'source_code': code_str,
                    'target_vocab': '',
                    'target_code': '',
                    'domain': '',
                    'status': 'proposed',
                    'origins': 'HT-FHIR',
                    'description': str(text or ''),
                    'occurrences': occurrences or 0,
                })

        # --- Sheet F: other clinical top 2000 ---
        if 'F_Other_clinical_top2000' in wb.sheetnames:
            ws = wb['F_Other_clinical_top2000']
            for row in ws.iter_rows(min_row=2, values_only=True):
                code, vocabulary, text, occurrences, resource_types = (
                    row[0], row[1], row[2], row[3], row[4],
                )
                if code is None or vocabulary is None:
                    continue
                stats['F_total'] += 1
                vocab_id = str(vocabulary).strip()
                code_str = str(code).strip()
                key = (vocab_id, code_str)
                if key in existing_keys or key in seen:
                    stats['F_overlap'] += 1
                    continue
                seen.add(key)
                stats['F_added'] += 1
                rows.append({
                    'source_vocab': vocab_id,
                    'source_code': code_str,
                    'target_vocab': '',
                    'target_code': '',
                    'domain': '',
                    'status': 'proposed',
                    'origins': 'HT-FHIR',
                    'description': str(text or ''),
                    'occurrences': occurrences or 0,
                })

        wb.close()

        # Sort by occurrences descending for the output
        rows.sort(key=lambda r: -(r['occurrences'] or 0))

        # Report
        for prefix in ('A', 'B', 'C', 'F'):
            total = stats[f'{prefix}_total']
            overlap = stats[f'{prefix}_overlap']
            added = stats[f'{prefix}_added']
            self.stdout.write(f'  Sheet {prefix}: {total:,} total, {overlap:,} overlap, {added:,} net new')
        self.stdout.write(self.style.SUCCESS(f'Total rows for output: {len(rows):,}'))

        if options['dry_run']:
            self.stdout.write('Dry run — no file written.')
            return

        # Write Markdown
        output_path = Path(options['output'])
        with open(output_path, 'w') as f:
            f.write('# CureHub FHIR Code-to-Concept Mapping\n\n')
            f.write('Generated by `manage.py build_curehub_fhir_crossmap`; do not edit by hand.\n\n')
            f.write(f'Source: CureHub OMOP Mapping Worklist. {len(rows):,} non-overlapping rows.\n\n')
            f.write(
                '| Source system | Source code | Target OMOP vocabulary '
                '| Target OMOP code | Domain | Status | Origins '
                '| Candidate targets |\n'
            )
            f.write('| --- | --- | --- | --- | --- | --- | --- | --- |\n')
            for row in rows:
                f.write(
                    f'| {_escape_pipe(row["source_vocab"])} '
                    f'| {_escape_pipe(row["source_code"])} '
                    f'| {_escape_pipe(row["target_vocab"])} '
                    f'| {_escape_pipe(row["target_code"])} '
                    f'| {_escape_pipe(row["domain"])} '
                    f'| {row["status"]} '
                    f'| {row["origins"]} '
                    f'| {row["occurrences"]} |\n'
                )
        self.stdout.write(self.style.SUCCESS(f'Wrote {output_path}'))
